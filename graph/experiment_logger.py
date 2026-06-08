"""
ExperimentLogger — collects per-agent experiment data and writes a single-row
XLSX per execution, with all agents side-by-side in the compact format.

Column order (one row = one full execution):
    1. Header: Versione, Problema, Trattamento, # esecuzione (replica)
    2. For each agent N:
         Agent N - Iter 0 … Agent N - Iter MAX_REFINEMENTS
    3. Global social columns:
         # Test Totali (social), # Test Unici (social), Test Unici Passati (social)
    4. For each agent N (reordered by SF index then agent):
         Agent N - Social Feedback 0 … Agent N - Social Feedback MAX_CROSS_REFINEMENTS
    5. For each agent N:
         Agent N - Summary  (replaces Best Iteration / MI / MS / Tempo columns)

Iter cell format:      "passed/total - coverage%"  e.g. "26/27 - 100,0%"
Social Feedback format: "passed/total - coverage% - MI:xx,xx - MS:xx,xx"
Summary format:         "#rank - MI:xx,xx - MS:xx,xx - T:xx,xxs"

Styling:
  - Header row: bold + grey fill + wrap text
  - Best Iter cell per agent: bold
  - Best Social Feedback cell per SF index (highest tests passed): bold
  - Column widths: auto-sized, capped at 60
"""
import csv
import logging
import os
import re
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _dump_csv(xlsx_path: Path):
    """Mirror the XLSX active sheet to a CSV next to it."""
    csv_path = xlsx_path.with_suffix(".csv")
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
    logging.info(f"Experiment CSV written to {csv_path}")

MAX_REFINEMENTS = int(os.getenv("MAX_REFINEMENTS", "3"))
MAX_CROSS_REFINEMENTS = int(os.getenv("MAX_CROSS_REFINEMENTS", os.getenv("MAX_REFINEMENTS", "3")))

BOLD_FONT = Font(bold=True)
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")

GLOBAL_HEADER_COLUMNS = [
    "Versione",
    "Problema",
    "Trattamento",
    "# esecuzione (replica)",
]

SOCIAL_GLOBAL_COLUMNS = [
    "# Test Totali (social)",
    "# Test Unici (social)",
    "Test Unici Passati (social)",
]


# ── formatting helpers ───────────────────────────────────────────────────────

def _fmt(value, decimals: int) -> str:
    """Format a float using comma as decimal separator (Italian locale)."""
    if value is None:
        return ""
    return f"{round(value, decimals)}".replace(".", ",")


def _iter_cell(tests_passed: int, tests_total: int, coverage_pct: float) -> str:
    """Format an Iter cell: 'passed/total - coverage%'."""
    cov = _fmt(coverage_pct, 1)
    return f"{tests_passed}/{tests_total} - {cov}%"


def _sf_cell(tests_passed: int, tests_total: int, coverage_pct: float,
             mi: Optional[float], ms: Optional[float]) -> str:
    """Format a Social Feedback cell: 'passed/total - coverage% - MI:xx - MS:xx'."""
    base = _iter_cell(tests_passed, tests_total, coverage_pct)
    mi_str = f" - MI:{_fmt(mi, 2)}" if mi is not None else ""
    ms_str = f" - MS:{_fmt(ms, 2)}" if ms is not None else ""
    return f"{base}{mi_str}{ms_str}"


# ── data class ───────────────────────────────────────────────────────────────

class IterationData:
    def __init__(
        self,
        tests_generated: int,
        tests_passed: int,
        tests_total: int,
        coverage_pct: float,
        maintainability_index: Optional[float],
        mutation_score: Optional[float] = None,
    ):
        self.tests_generated = tests_generated
        self.tests_passed = tests_passed
        self.tests_total = tests_total
        self.coverage_pct = coverage_pct
        self.maintainability_index = maintainability_index
        self.mutation_score = mutation_score


class AgentData:
    def __init__(
        self,
        agent_id: int,
        agent_name: str,
        agent_temp: str,
        iterations: list[IterationData],
        best_iteration_index: int,
        cross_test_iterations: list[IterationData],
        best_cross_test_iteration_index: int,
        is_best_model: bool,
        total_input_tokens: int,
        total_output_tokens: int,
        elapsed_time_seconds: float,
    ):
        self.agent_id = agent_id
        self.agent_name = agent_name
        self.agent_temp = agent_temp
        self.iterations = iterations
        self.best_iteration_index = best_iteration_index
        self.cross_test_iterations = cross_test_iterations
        self.best_cross_test_iteration_index = best_cross_test_iteration_index
        self.is_best_model = is_best_model
        self.total_input_tokens = total_input_tokens
        self.total_output_tokens = total_output_tokens
        self.elapsed_time_seconds = elapsed_time_seconds


# ── column builder ────────────────────────────────────────────────────────────

def _build_columns(num_agents: int) -> list[str]:
    cols = list(GLOBAL_HEADER_COLUMNS)

    # Iter columns per agent
    for n in range(num_agents):
        for i in range(MAX_REFINEMENTS + 1):
            cols.append(f"Agent {n} - Iter {i}")

    # Global social columns
    cols.extend(SOCIAL_GLOBAL_COLUMNS)

    # Social Feedback columns: ordered by SF index then agent
    for sf_i in range(MAX_CROSS_REFINEMENTS + 1):
        for n in range(num_agents):
            cols.append(f"Agent {n} - Social Feedback {sf_i}")

    # Single winner column
    cols.append("Winner")

    return cols


# ── logger ────────────────────────────────────────────────────────────────────

class ExperimentLogger:
    def __init__(self, output_path: Path):
        self.output_path = output_path

    def write(
        self,
        *,
        feedback_version: str,
        replica: int,
        problem: str,
        treatment: str,
        agents: list[AgentData],
        social_total_tests: Optional[int],
        social_unique_tests: Optional[int],
        social_unique_passed: Optional[int],
    ):
        num_agents = len(agents)
        columns = _build_columns(num_agents)
        col_index = {c: i + 1 for i, c in enumerate(columns)}

        # Build the row dict
        row: dict[str, object] = {}
        row["Versione"] = feedback_version
        row["Problema"] = problem
        row["Trattamento"] = treatment
        row["# esecuzione (replica)"] = replica

        for a in agents:
            n = a.agent_id
            for i in range(MAX_REFINEMENTS + 1):
                key = f"Agent {n} - Iter {i}"
                if i < len(a.iterations):
                    it = a.iterations[i]
                    row[key] = _iter_cell(it.tests_passed, it.tests_total, it.coverage_pct)
                else:
                    row[key] = ""

        row["# Test Totali (social)"] = social_total_tests if social_total_tests is not None else ""
        row["# Test Unici (social)"] = social_unique_tests if social_unique_tests is not None else ""
        row["Test Unici Passati (social)"] = social_unique_passed if social_unique_passed is not None else ""

        for sf_i in range(MAX_CROSS_REFINEMENTS + 1):
            for a in agents:
                n = a.agent_id
                key = f"Agent {n} - Social Feedback {sf_i}"
                if sf_i < len(a.cross_test_iterations):
                    it = a.cross_test_iterations[sf_i]
                    row[key] = _sf_cell(
                        it.tests_passed, it.tests_total, it.coverage_pct,
                        it.maintainability_index, it.mutation_score,
                    )
                else:
                    row[key] = ""

        # Winner column: single cell describing which agent/iteration won
        for a in agents:
            if a.is_best_model:
                n = a.agent_id
                bci = a.best_cross_test_iteration_index
                bi = a.best_iteration_index
                mi_val = ""
                ms_val = ""
                source = ""
                if bci >= 0 and bci < len(a.cross_test_iterations):
                    it = a.cross_test_iterations[bci]
                    mi_val = _fmt(it.maintainability_index, 2)
                    ms_val = _fmt(it.mutation_score, 2) if it.mutation_score is not None else ""
                    source = f"SF#{bci}"
                elif 0 <= bi < len(a.iterations):
                    it = a.iterations[bi]
                    mi_val = _fmt(it.maintainability_index, 2)
                    ms_val = _fmt(it.mutation_score, 2) if it.mutation_score is not None else ""
                    source = f"Iter#{bi}"
                t_val = _fmt(a.elapsed_time_seconds, 2)
                row["Winner"] = f"Agent {n} - {source} - MI:{mi_val} - MS:{ms_val} - T:{t_val}s"
                break

        # Determine bold cells
        bold_cols: set[str] = set()
        bold_cols.add("Winner")

        for a in agents:
            n = a.agent_id
            bi = a.best_iteration_index
            if bi >= 0:
                bold_cols.add(f"Agent {n} - Iter {bi}")

        for sf_i in range(MAX_CROSS_REFINEMENTS + 1):
            best_col = None
            best_passed = -1
            for a in agents:
                n = a.agent_id
                col_name = f"Agent {n} - Social Feedback {sf_i}"
                val = row.get(col_name, "")
                if val:
                    m = re.match(r"(\d+)/", str(val))
                    if m:
                        passed = int(m.group(1))
                        if passed > best_passed:
                            best_passed = passed
                            best_col = col_name
            if best_col:
                bold_cols.add(best_col)

        # Load or create workbook
        if self.output_path.exists():
            wb = load_workbook(self.output_path)
            ws = wb.active
            # Verify columns match; if not, recreate sheet
            existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            if existing_headers != columns:
                wb.remove(ws)
                ws = wb.create_sheet("Results", 0)
                self._write_header(ws, columns)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            self._write_header(ws, columns)

        # Append data row
        data_row_idx = ws.max_row + 1
        for field in columns:
            col_idx = col_index[field]
            value = row.get(field, "")
            cell = ws.cell(row=data_row_idx, column=col_idx, value=value)
            if field in bold_cols:
                cell.font = BOLD_FONT

        # Auto-size columns
        for col_idx, field in enumerate(columns, start=1):
            max_len = len(field)
            for row_idx in range(1, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value or ""
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logging.info(f"Experiment XLSX written to {self.output_path}")
        _dump_csv(self.output_path)

    def write_error(
        self,
        *,
        feedback_version: str,
        replica: int,
        problem: str,
        treatment: str,
        error_message: str,
        num_agents: int = 1,
    ):
        """Append an error row: header cols filled, iter cols empty, Winner = ERROR: msg."""
        columns = _build_columns(num_agents)
        col_index = {c: i + 1 for i, c in enumerate(columns)}
        row: dict[str, object] = {
            "Versione": feedback_version,
            "Problema": problem,
            "Trattamento": treatment,
            "# esecuzione (replica)": replica,
            "Winner": f"ERROR: {error_message}",
        }

        if self.output_path.exists():
            wb = load_workbook(self.output_path)
            ws = wb.active
            existing_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            if existing_headers != columns:
                wb.remove(ws)
                ws = wb.create_sheet("Results", 0)
                self._write_header(ws, columns)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            self._write_header(ws, columns)

        data_row_idx = ws.max_row + 1
        for field in columns:
            col_idx = col_index[field]
            value = row.get(field, "")
            cell = ws.cell(row=data_row_idx, column=col_idx, value=value)
            if field == "Winner":
                cell.font = BOLD_FONT

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.output_path)
        logging.info(f"Experiment XLSX error row written to {self.output_path}")

    def _write_header(self, ws, columns: list[str]):
        for col_idx, field in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True)
