"""Load Participium requirements from PDF and XLSX documents."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from thesis_rest_tester.domain.schemas import RequirementsCorpus, SourceRequirement


class RequirementsLoader:
    def load(
        self,
        description_pdf: str | Path,
        user_stories_xlsx: str | Path,
        faq_pdf: str | Path,
    ) -> RequirementsCorpus:
        paths = {
            "description PDF": Path(description_pdf),
            "user stories XLSX": Path(user_stories_xlsx),
            "FAQ PDF": Path(faq_pdf),
        }
        missing = [f"{label}: {path}" for label, path in paths.items() if not path.is_file()]
        if missing:
            raise FileNotFoundError("Missing requirements input file(s): " + "; ".join(missing))

        description_text = self._read_pdf(paths["description PDF"], "description PDF")
        faq_text = self._read_pdf(paths["FAQ PDF"], "FAQ PDF")
        user_stories = self._read_xlsx(paths["user stories XLSX"])
        source_requirements = self._source_requirements(user_stories)
        compact_text = self._compact(description_text, faq_text, user_stories)
        return RequirementsCorpus(
            description_text=description_text,
            faq_text=faq_text,
            user_stories=user_stories,
            source_requirements=source_requirements,
            compact_text=compact_text,
        )

    @staticmethod
    def _read_pdf(path: Path, label: str) -> str:
        try:
            reader = PdfReader(path)
            pages = [(page.extract_text() or "").strip() for page in reader.pages]
        except Exception as exc:
            raise ValueError(f"Could not read {label} at {path}: {exc}") from exc
        return "\n\n".join(page for page in pages if page)

    @staticmethod
    def _read_xlsx(path: Path) -> list[dict[str, Any]]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()
        except Exception as exc:
            raise ValueError(f"Could not read user stories XLSX at {path}: {exc}") from exc

        first_data_row = next(
            (index for index, row in enumerate(rows) if any(v is not None for v in row)),
            None,
        )
        if first_data_row is None:
            return []

        headers = RequirementsLoader._headers(rows[first_data_row])
        stories: list[dict[str, Any]] = []
        for row in rows[first_data_row + 1 :]:
            if not any(value is not None for value in row):
                continue
            padded = (*row, *([None] * max(0, len(headers) - len(row))))
            stories.append(
                {
                    header: RequirementsLoader._json_safe(value)
                    for header, value in zip(headers, padded, strict=False)
                }
            )
        return stories

    @staticmethod
    def _headers(row: tuple[Any, ...]) -> list[str]:
        headers: list[str] = []
        seen: dict[str, int] = {}
        for index, value in enumerate(row, start=1):
            base = (
                str(value).strip()
                if value is not None and str(value).strip()
                else f"column_{index}"
            )
            seen[base] = seen.get(base, 0) + 1
            headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        return headers

    @staticmethod
    def _json_safe(value: Any) -> Any:
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        return value

    @staticmethod
    def _source_requirements(user_stories: list[dict[str, Any]]) -> list[SourceRequirement]:
        requirements: list[SourceRequirement] = []
        seen_ids: set[str] = set()
        for row_number, row in enumerate(user_stories, start=1):
            normalized = {
                re.sub(r"[^a-z0-9]", "", key.lower()): value for key, value in row.items()
            }
            requirement_id = RequirementsLoader._first_value(
                normalized,
                "issueid",
                "requirementid",
                "id",
            )
            text = RequirementsLoader._first_value(
                normalized,
                "description",
                "story",
                "text",
            )
            if requirement_id is None or text is None:
                continue
            identifier = str(requirement_id).strip()
            if not identifier or identifier in seen_ids:
                raise ValueError(
                    "Duplicate or blank requirement ID in XLSX row "
                    f"{row_number + 1}: {identifier!r}"
                )
            seen_ids.add(identifier)

            text_value = str(text).strip()
            explicit_role = RequirementsLoader._first_value(normalized, "role", "actor")
            role = str(explicit_role).strip() if explicit_role is not None else "unspecified"
            role_match = re.search(r"^\s*As\s+an?\s+(.+?)(?:\n|\r|,?\s+I\s+want)", text_value, re.I)
            if explicit_role is None and role_match:
                role = role_match.group(1).strip()

            expected_behaviors: list[str] = []
            behavior_match = re.search(r"So\s+that\s+(.+?)(?:\n|\r|$)", text_value, re.I)
            if behavior_match:
                expected_behaviors.append(behavior_match.group(1).strip().rstrip("."))

            comments = RequirementsLoader._first_value(
                normalized,
                "commentscheckalsosystemtextualdescription",
                "comments",
                "constraints",
            )
            constraints = []
            if comments is not None and str(comments).strip():
                constraints.append(str(comments).strip())
            requirements.append(
                SourceRequirement(
                    id=identifier,
                    source="user_stories_xlsx",
                    text=text_value,
                    role=role,
                    business_value=RequirementsLoader._first_value(
                        normalized,
                        "businessvalue",
                        "priority",
                    ),
                    constraints=constraints,
                    expected_behaviors=expected_behaviors,
                )
            )
        return requirements

    @staticmethod
    def _first_value(normalized_row: dict[str, Any], *keys: str) -> Any:
        for key in keys:
            value = normalized_row.get(key)
            if value is not None:
                return value
        return None

    @staticmethod
    def _compact(
        description_text: str,
        faq_text: str,
        user_stories: list[dict[str, Any]],
    ) -> str:
        story_lines = []
        for index, story in enumerate(user_stories, start=1):
            fields = " | ".join(
                f"{key}: {value}" for key, value in story.items() if value is not None
            )
            story_lines.append(f"{index}. {fields}")
        stories_text = "\n".join(story_lines) or "(no user stories found)"
        return (
            "# Participium description\n"
            f"{description_text or '(no extractable text)'}\n\n"
            "# Participium user stories\n"
            f"{stories_text}\n\n"
            "# Participium FAQ\n"
            f"{faq_text or '(no extractable text)'}\n"
        )
